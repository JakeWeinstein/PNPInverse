#!/usr/bin/env python
"""Phase 7.2 Stage 0 — extract the K2SO4 pH 6.39 dual-series RRDE target.

Source: data/EChem Reactor Modeling-Seitz-Mangan/Brianna/
        '0,1M K2SO4 data 8-15-19.xlsx'  (Brianna Ruggiero, 2019-08-15)

Per the session-43-approved plan (~/.claude/plans/
phase7p2-k2so4-dual-series-fit.md):

- Parse RAW `cycle 2`/`cycle 3` sheets (cycle 1 excluded: in-sheet
  remark "pH changed to 6.11 while ring cleaning" + grossly different
  trace).  Re-derive every processed column from raw + in-sheet
  constants and cross-check against the workbook's own cached values
  with |delta| <= max(1e-8 abs, 1e-6 rel); against the (rounded)
  main-sheet copy of cycle 2 with 1e-4 abs.
- iR audit: the sheet computes  V_sheet = V_cal + (I_mA/1000)*Rs
  (their "iR New" column).  With the file's own anodic-positive
  current convention the PHYSICAL correction is
  V_phys = V_cal - (I_mA/1000)*Rs  (interface potential = measured
  minus ohmic drop).  Both axes are emitted; the canonical fit axis
  and the discrepancy are documented in the provenance note.
- Ring baseline: per-cycle median of raw j_ring over the pre-ORR
  shelf V_phys in [0.60, 1.00] (ring is flat there; the original
  disk-non-cathodic rule caught only 11 scan-start transient
  points), subtracted; spread goes into sigma.
- Disk background (DOCUMENTED STAGE-0 DEVIATION): the data shows a
  ring-silent cathodic shelf of -0.21..-0.33 mA/cm2 over
  V_phys 0.6..1.0 — ~60x the Exp Info cap band, ~7% of the plateau;
  it is capacitive/pseudocapacitive LSV background (scan-rate
  artifact), not ORR, and a steady-state model must not fit it.
  Treatment: linear background fit anchored on V_phys in
  [0.65, 1.00], subtracted across the sweep; sigma_bg = sqrt(MAD^2
  + (0.5*|slope|*extrapolation_span)^2) added in quadrature to the
  disk sigma; up-sweep cross-check ((down+up)/2 cancels first-order
  capacitive current) reported as a residual-systematic diagnostic.
  A background-scale x{0.5, 1.5} REFIT joins the Stage 4
  sensitivity table.
- Canonical RRDE quantities from CURRENTS (baseline-corrected ring):
      Sel% = 200*(I_r/N) / (|I_d| + I_r/N)
      n_e  = 4*|I_d| / (|I_d| + I_r/N)
  The sheet's own Sel%/n_e mix densities on different areas
  (overweights the ring term by A_d/A_r = 1.786) and are provenance
  only.
- Ring -> disk-equivalent peroxide current (PLOTS ONLY, the
  objective fits raw ring):
      pc = -j_ring * A_r / (N * A_d)   [mA/cm^2_disk, cathodic-neg]
      (= -j_ring * 0.109956 / (0.224 * 0.196350))
- Binning: ~30 bins over the predeclared fit window; per-bin mean of
  cycles 2+3; sigma = max(|c2-c3|/sqrt(2), per-series floor,
  0.02*|j|).  sigma is a conservative single-observation predictive
  scale (NOT the SEM of the mean).
- Fit window: lower edge = common data start; upper edge = largest V
  where |j_disk| > max(3*sigma_floor_disk, 2*cap_band) sustained for
  5 consecutive raw points.
- Air-saturation impossibility check + provenance numbers emitted to
  JSON for the provenance note.

Outputs (git add -f per repo precedent):
  data/k2so4_ph6p39_rrde_cycle2.csv
  data/k2so4_ph6p39_rrde_cycle3.csv
  data/k2so4_ph6p39_rrde_binned.csv
  StudyResults/phase7p2_stage0_extraction/{qa_overlay.png,
      sel_ne_qa.png, extraction_report.json}

Run from PNPInverse/ with the firedrake venv python (numpy/openpyxl/
matplotlib only — no Firedrake).
"""

from __future__ import annotations

import json
import os
from pathlib import Path

import numpy as np
import openpyxl

XLSX = ("data/EChem Reactor Modeling-Seitz-Mangan/Brianna/"
        "0,1M K2SO4 data 8-15-19.xlsx")
OUT_DATA = Path("data")
OUT_DIR = Path("StudyResults/phase7p2_stage0_extraction")

N_BINS = 30
F_MODEL = 0.02          # fractional model-error sigma floor
SUSTAIN = 5             # consecutive raw points for the window edge
TOL_ABS = 1e-8
TOL_REL = 1e-6
TOL_MAIN_ABS = 1e-4     # main sheet stores cycle-2 rounded to 5 dp

# O2 transport ceilings for the air-saturation check (mA/cm^2):
#   j_lim(4e) = n F D c / L ; production value 5.71 at c=1.2 mol/m3,
#   L=15.4 um (Phase 7 constants).  Scales as c/L.
J4E_CEILING_15P4 = 5.71
C_O2_PROD = 1.2
C_O2_AIR = 0.25
L_BRACKET_UM = (12.0, 15.4, 21.7)


def load_cycle(wb, name):
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    const = {
        "a_disk_cm2": float(rows[0][4]),
        "a_ring_cm2": float(rows[1][4]),
        "rs_ohm": float(rows[3][1]),
        "n_coll": float(rows[8][3]),
        "ref_cal_v": float(rows[8][4]),
    }
    cols = {k: [] for k in
            ("e_raw", "i_disk_ma", "i_ring_ma", "v_cal", "v_sheet",
             "j_disk", "j_ring", "sel_sheet", "ne_sheet")}
    idx = {"e_raw": 0, "i_disk_ma": 1, "i_ring_ma": 2, "v_cal": 5,
           "v_sheet": 6, "j_disk": 7, "j_ring": 8, "sel_sheet": 9,
           "ne_sheet": 10}
    for r in rows[8:]:
        if r[0] is None or not isinstance(r[0], (int, float)):
            continue
        for k, j in idx.items():
            v = r[j]
            cols[k].append(float(v) if isinstance(v, (int, float))
                           else np.nan)
    return const, {k: np.asarray(v) for k, v in cols.items()}


def crosscheck_rederive(name, const, arr):
    """Re-derive processed columns from raw + constants; compare to
    the workbook's cached values."""
    failures = []
    v_cal = arr["e_raw"] - const["ref_cal_v"]
    v_sheet = v_cal + (arr["i_disk_ma"] / 1000.0) * const["rs_ohm"]
    j_disk = arr["i_disk_ma"] / const["a_disk_cm2"]
    j_ring = arr["i_ring_ma"] / const["a_ring_cm2"]
    sel_sheet = 200.0 * j_ring / (const["n_coll"] * np.abs(j_disk)
                                  + j_ring)
    ne_sheet = 4.0 * np.abs(j_disk) / (np.abs(j_disk)
                                       + j_ring / const["n_coll"])
    for label, ours, cached in (
            ("v_cal", v_cal, arr["v_cal"]),
            ("v_sheet", v_sheet, arr["v_sheet"]),
            ("j_disk", j_disk, arr["j_disk"]),
            ("j_ring", j_ring, arr["j_ring"]),
            ("sel_sheet", sel_sheet, arr["sel_sheet"]),
            ("ne_sheet", ne_sheet, arr["ne_sheet"])):
        ok = np.isfinite(cached)
        d = np.abs(ours[ok] - cached[ok])
        tol = np.maximum(TOL_ABS, TOL_REL * np.abs(cached[ok]))
        bad = int(np.sum(d > tol))
        if bad:
            kmax = int(np.argmax(d - tol))
            failures.append(
                f"{name}/{label}: {bad} pts out of tol; worst "
                f"|d|={d[kmax]:.3e} at idx {kmax}")
    return {"v_cal": v_cal, "v_sheet": v_sheet, "j_disk": j_disk,
            "j_ring": j_ring}, failures


def crosscheck_main(wb, arr2_cached):
    """Main-sheet group 2 == cycle 2 processed (rounded to 5 dp)."""
    ws = wb["0.1 M K2SO4 aug 15"]
    rows = list(ws.iter_rows(min_row=4, values_only=True))
    v, jd, jr = [], [], []
    for r in rows:
        if isinstance(r[5], (int, float)):
            v.append(float(r[5])); jd.append(float(r[6]))
            jr.append(float(r[7]))
    n = min(len(v), len(arr2_cached["v_sheet"]))
    fails = []
    for label, main, cyc in (
            ("V", v, arr2_cached["v_sheet"]),
            ("j_disk", jd, arr2_cached["j_disk"]),
            ("j_ring", jr, arr2_cached["j_ring"])):
        d = np.abs(np.asarray(main[:n]) - np.asarray(cyc[:n]))
        bad = int(np.sum(d > TOL_MAIN_ABS))
        if bad:
            fails.append(f"main/{label}: {bad} pts beyond "
                         f"{TOL_MAIN_ABS} (max {d.max():.2e})")
    return n, fails


def sweep_segments(v):
    """Indices of the longest monotone-descending segment."""
    dv = np.diff(v)
    desc = dv < 0
    best, cur, best_len, cur_len = 0, 0, 0, 0
    start, best_start = 0, 0
    for i, d in enumerate(desc):
        if d:
            if cur_len == 0:
                start = i
            cur_len += 1
            if cur_len > best_len:
                best_len, best_start = cur_len, start
        else:
            cur_len = 0
    return best_start, best_start + best_len + 1


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

    report = {"source": XLSX, "tolerances": {
        "rederive": f"max({TOL_ABS} abs, {TOL_REL} rel)",
        "main_sheet": f"{TOL_MAIN_ABS} abs (5-dp rounded copy)"}}
    cycles = {}
    all_fails = []
    for name in ("cycle 1", "cycle 2", "cycle 3"):
        const, arr = load_cycle(wb, name)
        derived, fails = crosscheck_rederive(name, const, arr)
        all_fails += fails
        cycles[name] = (const, arr, derived)
        report.setdefault("constants", {})[name] = const
        report.setdefault("n_raw_rows", {})[name] = int(
            len(arr["e_raw"]))

    n_main, main_fails = crosscheck_main(wb, cycles["cycle 2"][1])
    all_fails += main_fails
    report["main_sheet_rows_checked"] = n_main
    report["crosscheck_failures"] = all_fails

    const = cycles["cycle 2"][0]
    a_d, a_r = const["a_disk_cm2"], const["a_ring_cm2"]
    n_coll, rs = const["n_coll"], const["rs_ohm"]

    # --- per-cycle canonical series (cycles 2+3 only) ---
    out_cycles = {}
    for name in ("cycle 2", "cycle 3"):
        c, arr, der = cycles[name]
        v_cal = der["v_cal"]
        i_d, i_r = arr["i_disk_ma"], arr["i_ring_ma"]
        v_sheet = der["v_sheet"]
        v_phys = v_cal - (i_d / 1000.0) * rs       # physical iR sign
        j_d = der["j_disk"]
        j_r = der["j_ring"]

        s0, s1 = sweep_segments(arr["e_raw"])
        sl = slice(s0, s1)
        vp, vs, vc = v_phys[sl], v_sheet[sl], v_cal[sl]
        jd_raw, jr_raw = j_d[sl], j_r[sl]
        idm = i_d[sl]

        # ring baseline: pre-ORR shelf (ring flat there)
        base_mask = (vp >= 0.60) & (vp <= 1.00)
        base = float(np.median(jr_raw[base_mask]))
        base_spread = float(np.std(jr_raw[base_mask]))
        jr_corr = jr_raw - base
        ir_corr = jr_corr * a_r

        # disk background: linear fit anchored on [0.65, 1.00]
        bg_mask = (vp >= 0.65) & (vp <= 1.00)
        slope, icpt = np.polyfit(vp[bg_mask], jd_raw[bg_mask], 1)
        bg = slope * vp + icpt
        bg_mad = float(np.median(np.abs(
            jd_raw[bg_mask] - bg[bg_mask])))
        extrap = np.maximum(0.0, 0.65 - vp)
        sig_bg = np.sqrt(bg_mad ** 2
                         + (0.5 * abs(slope) * extrap) ** 2)
        jd_corr = jd_raw - bg
        id_corr = jd_corr * a_d

        # up-sweep cross-check: (down+up)/2 cancels capacitive term
        up_v, up_j = v_phys[s1:], j_d[s1:]
        if len(up_v) > 50:
            order = np.argsort(up_v)
            up_on_down = np.interp(vp, up_v[order], up_j[order])
            avg_sweeps = 0.5 * (jd_raw + up_on_down)
            sweep_resid = jd_corr - avg_sweeps
        else:
            sweep_resid = np.full_like(vp, np.nan)

        with np.errstate(divide="ignore", invalid="ignore"):
            sel_canon = 200.0 * (ir_corr / n_coll) / (
                np.abs(id_corr) + ir_corr / n_coll)
            ne_canon = 4.0 * np.abs(id_corr) / (
                np.abs(id_corr) + ir_corr / n_coll)
        pc_equiv = -jr_corr * a_r / (n_coll * a_d)

        out_cycles[name] = dict(
            v_phys=vp, v_sheet=vs, v_cal=vc,
            j_disk_raw=jd_raw, j_disk=jd_corr, sig_bg=sig_bg,
            j_ring_raw=jr_raw, j_ring=jr_corr, i_disk_ma=idm,
            sel_canon=sel_canon, ne_canon=ne_canon,
            pc_equiv=pc_equiv, baseline=base,
            baseline_spread=base_spread,
            bg_slope=float(slope), bg_icpt=float(icpt),
            bg_mad=bg_mad, sweep_resid=sweep_resid,
            sweep=(int(s0), int(s1), int(len(v_cal))))
        report.setdefault("ring_baseline", {})[name] = {
            "median_mA_cm2_ring": base, "spread": base_spread,
            "n_pts": int(base_mask.sum()),
            "region_v_phys": [0.60, 1.00]}
        report.setdefault("disk_background", {})[name] = {
            "model": "linear, anchored V_phys [0.65, 1.00], "
                     "subtracted; capacitive LSV background "
                     "(ring-silent shelf), NOT ORR",
            "slope_mA_cm2_per_V": float(slope),
            "intercept_mA_cm2": float(icpt),
            "anchor_mad_mA_cm2": bg_mad,
            "bg_at_window_ends_mA_cm2": [
                float(bg[np.argmin(vp)]), float(bg[np.argmax(vp)])],
            "upsweep_crosscheck_median_abs_resid": float(
                np.nanmedian(np.abs(sweep_resid))),
            "upsweep_crosscheck_p90_abs_resid": float(
                np.nanpercentile(np.abs(sweep_resid), 90))}
        report.setdefault("sweep_segment", {})[name] = \
            out_cycles[name]["sweep"]

    # --- iR audit numbers ---
    c2 = out_cycles["cycle 2"]
    imax = int(np.argmax(np.abs(c2["j_disk"])))
    report["ir_audit"] = {
        "convention": "raw I anodic-positive (verified: +I at "
                      "V_RHE~1.09 pre-ORR, -I on ORR plateau)",
        "sheet_formula": "V_cal + (I_mA/1000)*Rs  (NONSTANDARD sign)",
        "physical_formula": "V_cal - (I_mA/1000)*Rs",
        "rs_ohm": rs,
        "max_axis_discrepancy_V": float(
            2 * np.abs(c2["i_disk_ma"][imax]) / 1000.0 * rs),
        "at_j_disk": float(c2["j_disk"][imax]),
    }

    # --- sigma floors ---
    # Exp Info capacitance band (+/-0.0008, assumed mA) is 60x
    # smaller than the observed LSV background shelf; the disk
    # floor is therefore set from the background-anchor MAD.
    cap_band_j = 0.0008 / a_d
    sig_floor_disk = max(cap_band_j,
                         out_cycles["cycle 2"]["bg_mad"],
                         out_cycles["cycle 3"]["bg_mad"])
    sig_floor_ring = max(
        out_cycles["cycle 2"]["baseline_spread"],
        out_cycles["cycle 3"]["baseline_spread"])
    report["sigma_floors"] = {
        "disk_mA_cm2": sig_floor_disk,
        "ring_mA_cm2_ring": sig_floor_ring,
        "cap_band_assumed_units": "mA (raw current) -> /A_d; "
                                  "demoted (background shelf 60x "
                                  "larger)",
        "f_model": F_MODEL}

    # --- fit window on the physical axis ---
    # background-corrected disk must clear its own uncertainty,
    # sustained, scanning from the anodic end of the down-sweep
    v2, jd2 = c2["v_phys"], c2["j_disk"]
    thresh_arr = 3 * np.maximum(c2["sig_bg"], sig_floor_disk)
    cath = jd2 < -thresh_arr
    run = 0
    v_up = None
    for k in range(len(v2)):
        run = run + 1 if cath[k] else 0
        if run >= SUSTAIN:
            v_up = float(v2[k - SUSTAIN + 1])
            break
    thresh = float(np.median(thresh_arr))
    v_lo = float(max(out_cycles["cycle 2"]["v_phys"].min(),
                     out_cycles["cycle 3"]["v_phys"].min()))
    report["fit_window"] = {
        "v_lo_phys": v_lo, "v_up_phys": v_up,
        "threshold_mA_cm2": thresh, "sustain_pts": SUSTAIN}

    # --- binning ---
    centers = np.linspace(v_lo, v_up, N_BINS)

    def interp(name, key):
        cy = out_cycles[name]
        order = np.argsort(cy["v_phys"])
        return np.interp(centers, cy["v_phys"][order],
                         cy[key][order])

    binned = {"v_phys": centers}
    for key, floor in (("j_disk", sig_floor_disk),
                       ("j_ring", sig_floor_ring)):
        a = interp("cycle 2", key)
        b = interp("cycle 3", key)
        mean = 0.5 * (a + b)
        sig = np.maximum.reduce([
            np.abs(a - b) / np.sqrt(2.0),
            np.full_like(mean, floor),
            F_MODEL * np.abs(mean)])
        binned[key] = mean
        binned[f"sigma_{key}"] = sig
    # background-subtraction systematic, in quadrature (disk)
    sig_bg_bins = 0.5 * (interp("cycle 2", "sig_bg")
                         + interp("cycle 3", "sig_bg"))
    binned["sigma_j_disk"] = np.sqrt(binned["sigma_j_disk"] ** 2
                                     + sig_bg_bins ** 2)
    binned["pc_equiv"] = -binned["j_ring"] * a_r / (n_coll * a_d)
    binned["sigma_pc_equiv"] = (binned["sigma_j_ring"] * a_r
                                / (n_coll * a_d))

    # n_e diagnostic band wherever the disk clears threshold
    i_d_bin = binned["j_disk"] * a_d
    i_r_bin = binned["j_ring"] * a_r
    mask = np.abs(binned["j_disk"]) > 10 * sig_floor_disk
    with np.errstate(divide="ignore", invalid="ignore"):
        ne = 4 * np.abs(i_d_bin) / (np.abs(i_d_bin)
                                    + i_r_bin / n_coll)
        dne_dir = -4 * np.abs(i_d_bin) / (
            np.abs(i_d_bin) + i_r_bin / n_coll) ** 2 / n_coll
        sig_ne = np.abs(dne_dir) * binned["sigma_j_ring"] * a_r
    binned["ne_canon"] = np.where(mask, ne, np.nan)
    binned["sigma_ne"] = np.where(mask, sig_ne, np.nan)
    lo = np.nanmin((binned["ne_canon"] - binned["sigma_ne"])[mask])
    hi = np.nanmax((binned["ne_canon"] + binned["sigma_ne"])[mask])
    report["ne_diagnostic"] = {
        "band_min": float(lo), "band_max": float(hi),
        "alarm": bool(lo < 1.8 or hi > 4.2),
        "n_eval_bins": int(mask.sum())}

    # --- air-saturation impossibility check ---
    plateau = float(np.abs(binned["j_disk"]).max())
    air = {}
    for L in L_BRACKET_UM:
        ceil_air = J4E_CEILING_15P4 * (15.4 / L) * (C_O2_AIR
                                                    / C_O2_PROD)
        air[f"L_{L}um"] = {
            "ceiling_air_mA_cm2": ceil_air,
            "excluded": bool(plateau > ceil_air)}
    report["air_saturation_check"] = {
        "measured_plateau_mA_cm2": plateau, **air}

    # --- canonical vs sheet selectivity reconciliation numbers ---
    k = int(np.argmax(out_cycles["cycle 2"]["sel_canon"]
                      * (np.abs(out_cycles["cycle 2"]["j_disk"])
                         > 10 * sig_floor_disk)))
    report["selectivity_reconciliation"] = {
        "note": "sheet formula mixes densities on different areas; "
                "overweights ring term by A_d/A_r = "
                f"{a_d / a_r:.4f}",
        "peak_sel_canonical_pct": float(np.nanmax(np.where(
            np.abs(out_cycles['cycle 2']['j_disk'])
            > 10 * sig_floor_disk,
            out_cycles['cycle 2']['sel_canon'], np.nan))),
        "exp_info_sheet_value_pct": 73.2}

    # --- write CSVs ---
    for name, fname in (("cycle 2", "k2so4_ph6p39_rrde_cycle2.csv"),
                        ("cycle 3", "k2so4_ph6p39_rrde_cycle3.csv")):
        cy = out_cycles[name]
        cols = ["v_phys", "v_sheet", "v_cal", "j_disk",
                "j_disk_raw", "sig_bg", "j_ring_raw",
                "j_ring", "pc_equiv", "sel_canon", "ne_canon"]
        hdr = ",".join(cols)
        meta = (f"# source={XLSX} sheet='{name}' (descending sweep "
                f"{cy['sweep'][0]}:{cy['sweep'][1]} of "
                f"{cy['sweep'][2]} rows)\n"
                f"# A_d={a_d} cm2, A_r={a_r} cm2, N={n_coll}, "
                f"Rs={rs} ohm, ref_cal={const['ref_cal_v']} V\n"
                f"# ring baseline {cy['baseline']:.6f} mA/cm2_ring "
                f"subtracted from j_ring (spread "
                f"{cy['baseline_spread']:.6f}, region V_phys "
                f"[0.60,1.00])\n"
                f"# disk LSV background bg(V) = {cy['bg_slope']:.5f}"
                f"*V + {cy['bg_icpt']:.5f} (anchored [0.65,1.00], "
                f"MAD {cy['bg_mad']:.5f}) subtracted: j_disk = "
                f"j_disk_raw - bg; sig_bg = extrapolation "
                f"uncertainty\n"
                f"# v_phys = V_cal - I*Rs (physical); v_sheet = "
                f"V_cal + I*Rs (workbook 'iR New', nonstandard "
                f"sign)\n"
                f"# j units mA/cm2 (disk: disk area, cathodic-neg; "
                f"ring: ring area, anodic-pos); pc_equiv = "
                f"-j_ring*A_r/(N*A_d) [mA/cm2_disk] PLOTS ONLY\n")
        with open(OUT_DATA / fname, "w") as f:
            f.write(meta + hdr + "\n")
            for i in range(len(cy["v_phys"])):
                f.write(",".join(f"{cy[c][i]:.8g}" for c in cols)
                        + "\n")

    bcols = ["v_phys", "j_disk", "sigma_j_disk", "j_ring",
             "sigma_j_ring", "pc_equiv", "sigma_pc_equiv",
             "ne_canon", "sigma_ne"]
    with open(OUT_DATA / "k2so4_ph6p39_rrde_binned.csv", "w") as f:
        f.write(f"# binned target, {N_BINS} bins on v_phys in "
                f"[{v_lo:.4f}, {v_up:.4f}] V_RHE; mean of cycles "
                f"2+3; sigma=max(|c2-c3|/sqrt2, floor, "
                f"{F_MODEL}*|j|); conservative single-obs scale\n")
        f.write(",".join(bcols) + "\n")
        for i in range(N_BINS):
            f.write(",".join(f"{binned[c][i]:.8g}" for c in bcols)
                    + "\n")

    with open(OUT_DIR / "extraction_report.json", "w") as f:
        json.dump(report, f, indent=1, default=str)

    # --- QA plots ---
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    fig, axes = plt.subplots(2, 2, figsize=(12, 8))
    ax = axes[0, 0]
    for name, color in (("cycle 2", "C0"), ("cycle 3", "C1")):
        cy = out_cycles[name]
        ax.plot(cy["v_sheet"], cy["j_disk"], color=color, lw=0.7,
                alpha=0.5, label=f"{name} (sheet axis)")
        ax.plot(cy["v_phys"], cy["j_disk"], color=color, lw=1.2,
                ls="--", label=f"{name} (phys axis)")
    ax.errorbar(binned["v_phys"], binned["j_disk"],
                yerr=binned["sigma_j_disk"], fmt="k.", ms=4,
                label="binned (phys)")
    ax.set_xlabel("V vs RHE"); ax.set_ylabel("j_disk mA/cm$^2$")
    ax.legend(fontsize=7); ax.set_title("disk: axes + binned")
    ax = axes[0, 1]
    for name, color in (("cycle 2", "C0"), ("cycle 3", "C1")):
        cy = out_cycles[name]
        ax.plot(cy["v_phys"], cy["j_ring_raw"], color=color, lw=0.7,
                alpha=0.4, label=f"{name} raw")
        ax.plot(cy["v_phys"], cy["j_ring"], color=color, lw=1.2,
                label=f"{name} baseline-corr")
    ax.errorbar(binned["v_phys"], binned["j_ring"],
                yerr=binned["sigma_j_ring"], fmt="k.", ms=4)
    ax.set_xlabel("V vs RHE (phys)")
    ax.set_ylabel("j_ring mA/cm$^2$_ring")
    ax.legend(fontsize=7); ax.set_title("ring: baseline correction")
    ax = axes[1, 0]
    cy = out_cycles["cycle 2"]
    m = np.abs(cy["j_disk"]) > 10 * sig_floor_disk
    ax.plot(cy["v_phys"][m], cy["sel_canon"][m], "C0",
            label="canonical (currents)")
    _, arr2, _ = cycles["cycle 2"]
    s0, s1, _ = cy["sweep"]
    ax.plot(cy["v_phys"][m],
            arr2["sel_sheet"][s0:s1][m], "C3--",
            label="sheet formula")
    ax.set_xlabel("V vs RHE (phys)"); ax.set_ylabel("Sel H2O2 %")
    ax.legend(fontsize=7)
    ax.set_title("selectivity: canonical vs sheet")
    ax = axes[1, 1]
    ax.errorbar(binned["v_phys"], binned["ne_canon"],
                yerr=binned["sigma_ne"], fmt="k.", ms=4)
    ax.axhline(2, color="gray", ls=":"); ax.axhline(4, color="gray",
                                                    ls=":")
    ax.set_xlabel("V vs RHE (phys)"); ax.set_ylabel("n_e canonical")
    ax.set_title("n_e diagnostic band")
    fig.tight_layout()
    fig.savefig(OUT_DIR / "qa_overlay.png", dpi=150)

    print(json.dumps(report, indent=1, default=str))
    print("\nGATE:",
          "PASS" if not all_fails and not report["ne_diagnostic"][
              "alarm"] else "CHECK FAILURES ABOVE")


if __name__ == "__main__":
    main()
