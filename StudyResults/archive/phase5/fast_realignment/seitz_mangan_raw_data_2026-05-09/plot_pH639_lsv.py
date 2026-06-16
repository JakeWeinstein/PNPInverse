"""Plot the K2SO4 pH 6.39 RRDE LSV traces from Brianna's 2019 xlsx.

Source: data/EChem Reactor Modeling-Seitz-Mangan/Brianna/0,1M K2SO4 data 8-15-19.xlsx
Sheet:  0.1 M K2SO4 aug 15  (paired Cycle 1 / Cycle 2 blocks)
"""
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import openpyxl as xl

REPO = Path(__file__).resolve().parents[2]
SRC = (
    REPO
    / "data"
    / "EChem Reactor Modeling-Seitz-Mangan"
    / "Brianna"
    / "0,1M K2SO4 data 8-15-19.xlsx"
)
OUT = Path(__file__).parent / "K2SO4_pH639_RRDE_LSV.png"


def load_block(sheet, col_offset):
    """Pull (E_RHE, j_disk, j_ring, H2O2_pct, eta) from rows 4..end at col_offset."""
    rows = []
    for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, values_only=True):
        chunk = row[col_offset : col_offset + 5]
        if all(isinstance(v, (int, float)) for v in chunk):
            rows.append(chunk)
    arr = np.array(rows, dtype=float)
    return arr[:, 0], arr[:, 1], arr[:, 2], arr[:, 3], arr[:, 4]


def main():
    wb = xl.load_workbook(SRC, read_only=True, data_only=True)
    sh = wb["0.1 M K2SO4 aug 15"]

    # Cycle 1 starts at col 0, Cycle 2 at col 5
    e1, jd1, jr1, h2o2_1, eta1 = load_block(sh, 0)
    e2, jd2, jr2, h2o2_2, eta2 = load_block(sh, 5)

    print(f"Cycle 1: {len(e1)} points; V_RHE in [{e1.min():.3f}, {e1.max():.3f}]")
    print(f"Cycle 2: {len(e2)} points; V_RHE in [{e2.min():.3f}, {e2.max():.3f}]")
    print(f"Cycle 1 max H2O2%: {h2o2_1.max():.2f} at V={e1[h2o2_1.argmax()]:.3f}")
    print(f"Cycle 2 max H2O2%: {h2o2_2.max():.2f} at V={e2[h2o2_2.argmax()]:.3f}")

    fig, axes = plt.subplots(1, 3, figsize=(15, 4.5))

    ax = axes[0]
    ax.plot(e1, jd1, "C0-", lw=1.2, label="cycle 1")
    ax.plot(e2, jd2, "C1-", lw=1.2, label="cycle 2", alpha=0.85)
    ax.axhline(0, color="0.6", lw=0.6)
    ax.set_xlabel("E_disk (V vs RHE)")
    ax.set_ylabel("j_disk (mA/cm$^2$)")
    ax.set_title("Disk current")
    ax.legend(loc="best", fontsize=9)
    ax.grid(alpha=0.3)
    ax.invert_xaxis()

    ax = axes[1]
    ax.plot(e1, jr1, "C0-", lw=1.2, label="cycle 1")
    ax.plot(e2, jr2, "C1-", lw=1.2, label="cycle 2", alpha=0.85)
    ax.axhline(0, color="0.6", lw=0.6)
    ax.set_xlabel("E_disk (V vs RHE)")
    ax.set_ylabel("j_ring (mA/cm$^2$)")
    ax.set_title("Ring current (H$_2$O$_2$ collection @ +1.2 V)")
    ax.legend(loc="best", fontsize=9)
    ax.grid(alpha=0.3)
    ax.invert_xaxis()

    ax = axes[2]
    ax.plot(e1, h2o2_1, "C0-", lw=1.2, label="cycle 1")
    ax.plot(e2, h2o2_2, "C1-", lw=1.2, label="cycle 2", alpha=0.85)
    ax.set_xlabel("E_disk (V vs RHE)")
    ax.set_ylabel("Sel H$_2$O$_2$ (%)")
    ax.set_title("Peroxide selectivity")
    ax.set_ylim(0, 100)
    ax.legend(loc="best", fontsize=9)
    ax.grid(alpha=0.3)
    ax.invert_xaxis()

    fig.suptitle(
        "Brianna 2019 — 0.1 M K$_2$SO$_4$, pH 6.39, RRDE (1600 rpm, N=0.224)",
        fontsize=12,
    )
    fig.tight_layout()
    fig.savefig(OUT, dpi=150, bbox_inches="tight")
    print(f"\nSaved: {OUT}")


if __name__ == "__main__":
    main()
